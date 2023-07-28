import os
import io
import re
import datetime
import json
import pickle
import google.oauth2.credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import Workbook
from docx import Document
import requests
import googleapiclient.http

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

def get_credentials():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)
    return creds

def get_files_from_folder(service, folder_id):
    files = []
    page_token = None
    while True:
        response = service.files().list(
            q=f"'{folder_id}' in parents",
            spaces='drive',
            fields='nextPageToken, files(id, name, createdTime, mimeType)',
            pageToken=page_token
        ).execute()
        files.extend(response.get('files', []))
        page_token = response.get('nextPageToken', None)
        if not page_token:
            break
    return sorted(files, key=lambda x: x['createdTime'])

def read_word_file(file_id, service, creds):
    file = service.files().get(fileId=file_id, fields='mimeType').execute()
    mime_type = file.get('mimeType', '')

    if mime_type == 'application/vnd.google-apps.document':
        request = service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        fh = io.BytesIO()
        downloader = googleapiclient.http.MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        return fh, 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    else:
        response = requests.get(f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media", headers={"Authorization": f"Bearer {creds.token}"})
        return io.BytesIO(response.content), 'text/plain'

def read_text_file(file_id, service):
    response = service.files().get_media(fileId=file_id)
    content = response.execute()
    return content

def read_excel_file(file_id, service):
    response = service.files().get_media(fileId=file_id)
    content = response.execute()
    return content

def main():
    creds = get_credentials()
    service = build('drive', 'v3', credentials=creds)

    folder_id = '1t0RijmVV0VumfL7fDlqGRK2CkLpb6sqv'  # Replace with your folder ID
    files = get_files_from_folder(service, folder_id)

    wb = Workbook()
    sheet = wb.active

    sheet.cell(1, 1, "File Name")
    sheet.cell(1, 2, "File Contents")

    row = 2
    for file in files:
        file_id = file['id']
        file_name = file['name']
        created_time = datetime.datetime.fromisoformat(file['createdTime'][:-1])  # Removing 'Z' from the timestamp

        if file['mimeType'] == 'application/vnd.google-apps.document':
            contents, mime_type = read_word_file(file_id, service, creds)
            doc = Document(contents) if mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' else None
            file_contents = "\n".join([p.text for p in doc.paragraphs]) if doc else contents.getvalue().decode('utf-8')
        elif file['mimeType'] == 'text/plain':
            contents = read_text_file(file_id, service)
            file_contents = contents.decode('utf-8')
        elif file['mimeType'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            contents = read_excel_file(file_id, service)
            file_contents = "Excel files are not supported in this script."
        else:
            print(f"Skipping file '{file_name}' with unsupported MIME type '{file['mimeType']}'")
            continue

        sheet.cell(row, 1, file_name)
        sheet.cell(row, 2, file_contents)
        row += 1

    output_file = "output.xlsx"
    wb.save(output_file)
    print(f"Data successfully written to '{output_file}'.")

if __name__ == "__main__":
    main()
